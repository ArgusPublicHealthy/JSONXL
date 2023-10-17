import json
import requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import schedule
import time
import openpyxl
from datetime import datetime

APPID = '90dce1e419669e37d545fe6e359410c7'
location = 'San Francisco, US'

def get_weather_data(location):
    url = f'https://api.openweathermap.org/data/2.5/weather?q={location}&APPID={APPID}'
    response = requests.get(url)
    response.raise_for_status()
    return response.json()


def send_weather_alert():
    weather_data = get_weather_data(location)
    current_weather = weather_data['weather'][0]['main']
    weather_description = weather_data['weather'][0]['description']
    
    # print(weather_data)
    # print(current_weather)
    # print(weather_description)
    
        #openpyl code here or use Popen to run openpyxl script

    wb = openpyxl.Workbook()
    ws = wb.active

    #store weather information in ws
    #ws['A1'] =  weather_data
    ws['A3'] =  current_weather
    ws['A5'] =  weather_description

    wb.save("FrostAlert.xlsx")

    if current_weather.lower() == 'clouds' and 'few clouds' in weather_description.lower():
        username = 'zubairanwarrr@outlook.com'
        password = 'Azaan12$'
        smtp_server = 'smtp-mail.outlook.com'
        smtp_port = 587

        smtp_connection = smtplib.SMTP(smtp_server, smtp_port)
        smtp_connection.starttls()
        smtp_connection.login(username, password)

        from_email = username
        to_email = 'Zubair.anwar@cdph.ca.gov'
        subject = 'Frost Alert: Move Your Plants Indoors'
        message = 'The weather forecast indicates possible frost. Consider moving your plants indoors.'

        msg = MIMEMultipart()
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(message, 'plain'))

        smtp_connection.sendmail(from_email, to_email, msg.as_string())
        smtp_connection.quit()

def my_task():
    print("This task runs every Tuesday at 3:14 PM.")
    send_weather_alert()
schedule.every().tuesday.at("15:14").do(my_task)



while True:
    schedule.run_pending()
    time.sleep(1)